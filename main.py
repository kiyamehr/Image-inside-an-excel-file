from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill

work_book = openpyxl.Workbook()
work_sheet = work_book.active

img_name = "images.jpg"
img = Image.open(img_name).convert("RGB")
pix = img.load()
width, height = img.size


# for image loop
for y in range(0, height, 10):
    for x in range(0, width, 10):
        r, g, b = pix[x, y]

        # : -> means "start formatting"
        # 02 -> means "make it at least 2 characters long, and if it’s shorter, pad with zeros"
        # X -> means "convert the number to uppercase hexadecimal"
        hex_color = "{:02X}{:02X}{:02X}".format(r, g, b)
        fill = PatternFill(
            start_color="FF" + hex_color, end_color="FF" + hex_color, fill_type="solid"
        )

        # row/columns = 'numbers', casue openpyxl knows that num of column should transefer to words\
        # so we can call it directly on out code
        # and the division 'y//10' must be equal to the steps we are taking in the for loop.\
        # and we use +1 cause excel cells start with 1 but python starts with 0
        work_sheet.cell(row=y // 10 + 1, column=x // 10 + 1).fill = fill
        
print(f"Finished Converting {img_name} inside and Excel file.")

work_book.save("image_excel.xlsx")
